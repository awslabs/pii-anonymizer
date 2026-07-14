# Copyright Amazon.com, Inc. or its affiliates. All Rights Reserved.
# SPDX-License-Identifier: Apache-2.0

"""
Excel/CSV PII Processor Module

Handles PII detection and redaction in Excel (.xlsx) and CSV (.csv) files.
Uses 3-step pipeline:
1. Detect PII per sheet/chunk
2. Batch generate synthetic replacements (consistent across sheets)
3. Replace using unified mapping
"""

import csv
import io
import os
import tempfile
import logging
import json

import openpyxl

from helpers.model_config_helper import (
    get_inference_config_from_yaml,
    get_concurrency_config,
)
from core.synthetic_pii_generator import batch_generate_synthetic_pii
from helpers.threaded_detector import run_threaded_pii_detection
from helpers.text_chunker import chunk_text_by_lines
from core.text_replacer import replace_pii_in_text
from helpers.token_tracker import TokenTracker

logger = logging.getLogger(__name__)


def _token_bound_chunks(raw_chunks, cc):
    """Sub-split any structural chunk that exceeds the model's token limit.

    Excel/CSV split by sheet, but a single sheet can hold thousands of rows —
    enough PII that the detection OUTPUT exceeds a model's output cap (e.g. Nova
    tops out at 10000) and truncates, silently losing detections. This splits
    each oversized sheet at row (line) boundaries using the model-aware
    max_txt_chunk_tokens, so each piece's output stays within the model limit.
    Small sheets are left as a single chunk (no behavior change).
    """
    bounded = {}
    for name, text in raw_chunks.items():
        parts = chunk_text_by_lines(
            text, cc["max_txt_chunk_tokens"], cc["chars_per_token"]
        )
        if len(parts) <= 1:
            bounded[name] = text
        else:
            for i, part in enumerate(parts, 1):
                bounded[f"{name} (part {i})"] = part
            logger.info(
                f"Chunk '{name}' split into {len(parts)} parts "
                f"(exceeds {cc['max_txt_chunk_tokens']}-token model limit)"
            )
    return bounded


def detect_pii_excel(source_bucket, source_key, config, bedrock_runtime, s3_client):
    """Detect PII in an Excel file (Step 1 only)."""
    filename = os.path.splitext(os.path.basename(source_key))[0]
    model_id = config["model"]["id"]
    model_provider = config["model"]["provider"]
    inference_config = get_inference_config_from_yaml(config)
    cc = get_concurrency_config(config)
    tracker = TokenTracker(model_id)

    temp_fd, temp_path = tempfile.mkstemp(
        suffix=f"_{filename}.xlsx", dir=tempfile.gettempdir()
    )
    os.close(temp_fd)
    s3_client.download_file(source_bucket, source_key, temp_path)
    logger.info(f"Downloaded Excel file: {source_key}")

    from validation.document_validator import validate_excel, DocumentValidationError

    try:
        validate_excel(temp_path, config)
    except DocumentValidationError as e:
        os.remove(temp_path)
        raise ValueError(f"Excel validation failed: {e}")

    excel_data = extract_text_from_excel(temp_path)
    os.remove(temp_path)
    chunks = {}
    for sheet in excel_data["sheets"]:
        sheet_text = "\n".join([c["value"] for c in sheet["cells"]])
        if sheet_text.strip():
            chunks[sheet["sheet_name"]] = sheet_text
    chunks = _token_bound_chunks(chunks, cc)

    all_detections, failed_chunks, _ = run_threaded_pii_detection(
        chunks,
        model_id,
        model_provider,
        bedrock_runtime,
        inference_config,
        cc["max_workers"],
        label="Sheet",
        token_tracker=tracker,
        config=config,
    )

    return {
        "source_key": source_key,
        "file_type": "xlsx",
        "detections": [
            {
                "content": d["content"],
                "type": d["type"],
                "confidence": d.get("confidence", 0.0),
            }
            for d in all_detections
        ],
        "failed_chunks": failed_chunks,
        "token_usage": tracker.summary(),
    }


def extract_text_from_excel(xlsx_path):
    """Extract all text from an Excel workbook (all sheets)."""
    wb = openpyxl.load_workbook(xlsx_path)
    sheets_data = []
    all_text = []
    for sheet in wb.worksheets:
        cells = [
            {"sheet": sheet.title, "cell": c.coordinate, "value": str(c.value).strip()}
            for row in sheet.iter_rows()
            for c in row
            if c.value and isinstance(c.value, str) and str(c.value).strip()
        ]
        sheets_data.append({"sheet_name": sheet.title, "cells": cells})
        all_text.extend([c["value"] for c in cells])

    logger.info(f"Extracted text from {len(sheets_data)} sheets")
    return {
        "sheets": sheets_data,
        "all_text": "\n".join(all_text),
        "total_sheets": len(sheets_data),
    }


def extract_text_from_csv(csv_text):
    """Extract all text from CSV content, structured like Excel sheets."""
    rows = list(csv.reader(io.StringIO(csv_text)))
    cells = [
        {"sheet": "Sheet1", "cell": f"R{r + 1}C{c + 1}", "value": val.strip()}
        for r, row in enumerate(rows)
        for c, val in enumerate(row)
        if val and val.strip()
    ]
    logger.info(f"Extracted {len(cells)} cells from CSV")
    return {
        "sheets": [{"sheet_name": "Sheet1", "cells": cells}],
        "all_text": "\n".join(c["value"] for c in cells),
        "total_sheets": 1,
    }


def detect_pii_csv(source_bucket, source_key, config, bedrock_runtime, s3_client):
    """Detect PII in a CSV file (Step 1 only)."""
    model_id = config["model"]["id"]
    model_provider = config["model"]["provider"]
    inference_config = get_inference_config_from_yaml(config)
    cc = get_concurrency_config(config)
    tracker = TokenTracker(model_id)

    obj = s3_client.get_object(Bucket=source_bucket, Key=source_key)
    text = obj["Body"].read().decode("utf-8", errors="ignore")
    logger.info(f"Downloaded CSV file: {source_key}")

    csv_data = extract_text_from_csv(text)
    chunks = {}
    for sheet in csv_data["sheets"]:
        sheet_text = "\n".join([c["value"] for c in sheet["cells"]])
        if sheet_text.strip():
            chunks[sheet["sheet_name"]] = sheet_text
    chunks = _token_bound_chunks(chunks, cc)

    all_detections, failed_chunks, _ = run_threaded_pii_detection(
        chunks,
        model_id,
        model_provider,
        bedrock_runtime,
        inference_config,
        cc["max_workers"],
        label="Chunk",
        token_tracker=tracker,
        config=config,
    )

    return {
        "source_key": source_key,
        "file_type": "csv",
        "detections": [
            {
                "content": d["content"],
                "type": d["type"],
                "confidence": d.get("confidence", 0.0),
            }
            for d in all_detections
        ],
        "failed_chunks": failed_chunks,
        "token_usage": tracker.summary(),
    }


def extract_text_from_json(json_text):
    """Recursively extract all string values from JSON."""
    import json as json_mod

    data = json_mod.loads(json_text)
    values = []

    def _extract(obj, path=""):
        if isinstance(obj, str) and obj.strip():
            values.append({"path": path, "value": obj.strip()})
        elif isinstance(obj, dict):
            for k, v in obj.items():
                _extract(v, f"{path}.{k}" if path else k)
        elif isinstance(obj, list):
            for i, v in enumerate(obj):
                _extract(v, f"{path}[{i}]")

    _extract(data)
    logger.info(f"Extracted {len(values)} string values from JSON")
    return values


def detect_pii_json(source_bucket, source_key, config, bedrock_runtime, s3_client):
    """Detect PII in a JSON file (Step 1 only)."""
    model_id = config["model"]["id"]
    model_provider = config["model"]["provider"]
    inference_config = get_inference_config_from_yaml(config)
    cc = get_concurrency_config(config)
    tracker = TokenTracker(model_id)

    obj = s3_client.get_object(Bucket=source_bucket, Key=source_key)
    text = obj["Body"].read().decode("utf-8", errors="ignore")
    logger.info(f"Downloaded JSON file: {source_key}")

    values = extract_text_from_json(text)
    chunk_text = "\n".join(v["value"] for v in values)
    chunks = {"JSON": chunk_text} if chunk_text.strip() else {}

    all_detections, failed_chunks, _ = run_threaded_pii_detection(
        chunks,
        model_id,
        model_provider,
        bedrock_runtime,
        inference_config,
        cc["max_workers"],
        label="Chunk",
        token_tracker=tracker,
        config=config,
    )

    # Map detections back to JSON paths
    for d in all_detections:
        d["json_paths"] = [v["path"] for v in values if d["content"] in v["value"]]

    return {
        "source_key": source_key,
        "file_type": "json",
        "detections": [
            {
                "content": d["content"],
                "type": d["type"],
                "confidence": d.get("confidence", 0.0),
                "json_paths": d.get("json_paths", []),
            }
            for d in all_detections
        ],
        "failed_chunks": failed_chunks,
        "token_usage": tracker.summary(),
    }


def replace_pii_in_excel(xlsx_path, output_path, pii_mapping):
    """Replace PII in Excel workbook using exact, normalized, and fuzzy matching."""
    wb = openpyxl.load_workbook(xlsx_path)
    found_originals = set()
    replacements = 0
    all_match_types = {}
    all_text = []

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    all_text.append(str(cell.value))
                    new_val, cell_found, cell_count, cell_match_types, _ = (
                        replace_pii_in_text(str(cell.value), pii_mapping)
                    )
                    if cell_count > 0:
                        cell.value = new_val
                        found_originals.update(cell_found)
                        replacements += cell_count
                        all_match_types.update(cell_match_types)

    source_text = " ".join(all_text)
    not_found = len(pii_mapping) - len(found_originals)
    if not_found:
        missing = [orig for orig in pii_mapping if orig not in found_originals]
        covered = [m for m in missing if any(m in r for r in found_originals)]
        hallucinated = [m for m in missing if m not in covered and m not in source_text]
        truly_missing = len(missing) - len(covered) - len(hallucinated)
        if covered:
            logger.info(
                f"{len(covered)} PII skipped (already covered by longer replacement)"
            )
        if hallucinated:
            logger.info(
                f"{len(hallucinated)} PII skipped (LLM hallucinated, not in source text)"
            )
        if truly_missing:
            logger.warning(f"{truly_missing} PII not found in workbook")

    logger.info(
        f"Replacement: {len(found_originals)}/{len(pii_mapping)} unique PII replaced, {not_found} not found"
    )

    wb.save(output_path)
    return {
        "success": True,
        "replacements": replacements,
        "found_originals": found_originals,
    }


def _find_sheet_for_pii(pii_original, sheets_data):
    """Find which sheet a PII value appears in. Returns first match sheet name and number."""
    for idx, sheet in enumerate(sheets_data, 1):
        for cell in sheet["cells"]:
            if pii_original in cell["value"]:
                return sheet["sheet_name"], idx
    return "Unknown", 0


def store_excel_pii_mapping(
    detections,
    pii_mapping,
    filename,
    dynamodb_manager,
    sheets_data,
    found_originals,
    tracker=None,
):
    """Store Excel PII mappings in DynamoDB."""
    try:
        detailed_pii_data = []
        for det in detections:
            original = det["content"]
            synthetic = pii_mapping.get(original, "")
            if original in found_originals:
                status = "text_replaced"
            elif any(original in f for f in found_originals):
                status = "text_replaced"
                if not synthetic:
                    for f in found_originals:
                        if original in f and f in pii_mapping:
                            synthetic = f"(covered by: {pii_mapping[f]})"
                            break
            else:
                status = "not_redacted"
            record = {
                "original": original,
                "synthetic": synthetic,
                "type": det.get("type", "UNKNOWN"),
                "confidence": det.get("confidence", 0),
                "source": "excel",
                "replacement_status": status,
            }
            if status == "not_redacted":
                record["not_redacted_reason"] = "text_not_found_in_document"
            sheet_name, sheet_num = _find_sheet_for_pii(original, sheets_data)
            record["page_num"] = sheet_num
            record["sheet_name"] = sheet_name
            detailed_pii_data.append(record)

        token_usage = tracker.summary() if tracker else {}
        dynamodb_manager.store_pii_mapping(
            detailed_pii_data, filename, status="SUCCESS", token_usage=token_usage
        )
        if tracker:
            tracker.log_summary()
        replaced = sum(
            1 for d in detailed_pii_data if d["replacement_status"] == "text_replaced"
        )
        logger.info(
            f"Stored {len(detailed_pii_data)} Excel PII mappings in DynamoDB ({replaced} replaced, {len(detailed_pii_data) - replaced} not found)"
        )
        return True
    except Exception as e:
        logger.warning(f"Failed to store Excel PII mapping in DynamoDB: {e}")
        return False


def process_excel_file(
    source_bucket,
    source_key,
    output_bucket,
    filename_without_ext,
    config,
    bedrock_runtime,
    dynamodb_manager,
    s3_client,
    folder_path="",
):
    """End-to-end Excel file processing: detect per sheet, batch synthetic, replace."""
    try:
        model_id = config["model"]["id"]
        model_provider = config["model"]["provider"]
        inference_config = get_inference_config_from_yaml(config)

        # Download from S3
        temp_fd, temp_path = tempfile.mkstemp(
            suffix=f"_{filename_without_ext}.xlsx", dir=tempfile.gettempdir()
        )
        os.close(temp_fd)
        s3_client.download_file(source_bucket, source_key, temp_path)
        logger.info(f"Downloaded Excel file: {source_key}")

        # Validate Excel document
        from validation.document_validator import (
            validate_excel,
            DocumentValidationError,
        )

        try:
            validate_excel(temp_path, config)
            logger.info(f"Excel validation passed for: {filename_without_ext}")
        except DocumentValidationError as e:
            logger.error(f"Excel validation failed: {str(e)}")
            return {"success": False, "error": str(e), "pii_count": 0}

        # Step 1: Detect PII per sheet concurrently
        excel_data = extract_text_from_excel(temp_path)
        cc = get_concurrency_config(config)
        tracker = TokenTracker(model_id)
        chunks = {}
        for sheet in excel_data["sheets"]:
            sheet_text = "\n".join([c["value"] for c in sheet["cells"]])
            if sheet_text.strip():
                chunks[sheet["sheet_name"]] = sheet_text

        all_detections, failed_chunks, raw_detections = run_threaded_pii_detection(
            chunks,
            model_id,
            model_provider,
            bedrock_runtime,
            inference_config,
            cc["max_workers"],
            label="Sheet",
            token_tracker=tracker,
            config=config,
        )
        logger.info(
            f"Total detections across all sheets: {len(raw_detections)} raw, {len(all_detections)} after dedup"
        )

        if failed_chunks:
            logger.warning(f"Failed sheets: {[f['chunk_id'] for f in failed_chunks]}")
            if dynamodb_manager:
                for fc in failed_chunks:
                    dynamodb_manager.store_pii_mapping(
                        [],
                        filename_without_ext,
                        status="FAILED_DETECTION",
                        error_message=f"Sheet '{fc['chunk_id']}': {fc['error']}",
                    )

        # Step 2: Batch generate synthetic replacements
        pii_mapping = {}
        if all_detections:
            pii_mapping = batch_generate_synthetic_pii(
                all_detections,
                model_id,
                model_provider,
                bedrock_runtime,
                config=config,
                token_tracker=tracker,
            )
            logger.info(f"Generated {len(pii_mapping)} synthetic replacements")

        # Step 3: Replace using unified mapping
        fd, output_path = tempfile.mkstemp(
            suffix=f"_redacted_{filename_without_ext}.xlsx", dir=tempfile.gettempdir()
        )
        os.close(fd)
        result = replace_pii_in_excel(temp_path, output_path, pii_mapping)

        # Upload to S3
        s3_output_key = f"{folder_path}redacted_{filename_without_ext}.xlsx"
        s3_client.upload_file(output_path, output_bucket, s3_output_key)
        logger.info(f"Uploaded redacted Excel file to: {s3_output_key}")

        # Create and upload summary JSON
        summary = {
            "input_file": f"s3://{source_bucket}/{source_key}",
            "output_file": f"s3://{output_bucket}/{s3_output_key}",
            "pii_count": len(all_detections),
            "pii_replaced": len(result["found_originals"]),
            "pii_not_found": len(pii_mapping) - len(result["found_originals"]),
            "sheets_processed": len(excel_data["sheets"]),
            "token_usage": tracker.summary() if tracker else {},
        }
        summary_key = f"{folder_path}redaction_summary_{filename_without_ext}.json"
        s3_client.put_object(
            Body=json.dumps(summary, indent=2),
            Bucket=output_bucket,
            Key=summary_key,
            ContentType="application/json",
        )
        logger.info(f"Uploaded summary to: {summary_key}")

        # Store in DynamoDB
        if dynamodb_manager and all_detections:
            store_excel_pii_mapping(
                raw_detections,
                pii_mapping,
                filename_without_ext,
                dynamodb_manager,
                excel_data["sheets"],
                result["found_originals"],
                tracker,
            )

        # Cleanup
        os.remove(temp_path)
        os.remove(output_path)

        return {
            "success": True,
            "s3_output_file": s3_output_key,
            "summary_file": summary_key,
            "pii_count": len(all_detections),
            "replacements": result["replacements"],
        }
    except Exception as e:
        logger.error(f"Error processing Excel file: {e}", exc_info=True)
        if dynamodb_manager:
            dynamodb_manager.store_pii_mapping(
                [], filename_without_ext, status="FAILED", error_message=str(e)
            )
        return {"success": False, "error": str(e)}
